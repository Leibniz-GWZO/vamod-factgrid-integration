import pandas as pd
import re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def is_cell_yellow(cell):
    """Prüft ob eine Zelle gelb markiert ist"""
    if cell.fill and cell.fill.start_color:
        # Verschiedene Gelbtöne prüfen
        yellow_colors = ['FFFF00', 'FFFFFF00', 'FFFFE57F', 'FFFFFFFF00', 'FFFFEB9C', 'FFFFCC99']
        cell_color = str(cell.fill.start_color.index).upper()
        return any(yellow in cell_color for yellow in yellow_colors) or cell_color in yellow_colors
    return False

def aggregate_excel_rows(input_file, output_file):
    """
    Liest Excel-Datei und aggregiert verschiedene Gruppen von Zeilen
    Behält Formatierung bei
    """
    try:
        # Excel-Datei mit openpyxl laden (für Formatierung)
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Auch mit pandas laden (für einfachere Datenverarbeitung)
        df = pd.read_excel(input_file)
        print(f"Datei gelesen: {len(df)} Zeilen, {len(df.columns)} Spalten")
        
        # Prüfen ob die Spalte 'Name 1' existiert
        if 'Name 1' not in df.columns:
            print("Warnung: Spalte 'Name 1' nicht gefunden!")
            print("Verfügbare Spalten:", list(df.columns))
            return
        
        # Spaltenindex für 'Name 1' finden
        name1_col_idx = list(df.columns).index('Name 1') + 1  # openpyxl ist 1-basiert
        
        # Formatierung sammeln - prüfen welche Zeilen gelb markiert sind
        yellow_rows = set()
        yellow_cells = {}  # {(row, col): True}
        
        for row_idx in range(2, ws.max_row + 1):  # Ab Zeile 2 (Header überspringen)
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_cell_yellow(cell):
                    yellow_cells[(row_idx, col_idx)] = True
        
        # Spalten identifizieren, die nicht aggregiert werden sollen
        non_aggregated_columns = ['Name 1']  # Diese Spalten werden nie dupliziert
        
        # Alle anderen Spalten finden, die übernommen werden sollen
        static_columns = []
        numbered_patterns = defaultdict(list)
        
        for col in df.columns:
            if col in non_aggregated_columns:
                continue
            # Suche nach Mustern wie "Funktion 1", "Aussteller 2", etc.
            match = re.match(r'(.+?)\s+(\d+)$', col)
            if match:
                base_name = match.group(1)
                number = int(match.group(2))
                numbered_patterns[base_name].append((number, col))
            else:
                # Alle anderen Spalten sind statische Spalten (wie "Häufigkeit")
                static_columns.append(col)
        
        # Sortiere die nummerierten Spalten
        for base_name in numbered_patterns:
            numbered_patterns[base_name].sort()
        
        print(f"Statische Spalten (werden übernommen): {static_columns}")
        print(f"Nummerierte Spaltengruppen: {list(numbered_patterns.keys())}")
        
        # Definiere Aggregationsgruppen
        aggregation_groups = [
            ["JA"],  # JA-Zeilen
            ["KIII"],  # KIII-Zeilen
            ["WO"],  # WO-Zeilen
            ["LA"],  # LA-Zeilen
            ["MU"],  # MU-Zeilen
            ["Denharth"],  # Denharth-Zeilen
            ["Benko de Kuchar", "Benko de Zabokruki"],  # Benko-Gruppe
            ["Iesco de Mazouia", "Johannes Mazovita"],  # Mazovia-Gruppe
            ["Petrus Wlodkowicz", "Petrus Wlodkowicz de Charbinouicze"]  # Petrus-Gruppe
        ]
        
        print(f"Definiert: {len(aggregation_groups)} Aggregationsgruppen")
        
        # Zeilen ohne Aggregation sammeln
        result_df = df.copy()
        
        # Für jede Aggregationsgruppe
        for group_idx, group_values in enumerate(aggregation_groups):
            print(f"\n=== Verarbeite Gruppe {group_idx + 1}: {group_values} ===")
            
            # Zeilen dieser Gruppe finden
            group_mask = result_df['Name 1'].isin(group_values)
            group_rows = result_df[group_mask].copy()
            other_rows = result_df[~group_mask].copy()
            
            print(f"Gefunden: {len(group_rows)} Zeilen in Gruppe {group_values}")
            
            if len(group_rows) == 0:
                print(f"Keine Zeilen für Gruppe {group_values} gefunden, überspringe...")
                continue
            elif len(group_rows) == 1:
                print(f"Nur eine Zeile für Gruppe {group_values} gefunden, keine Aggregation nötig")
                continue
            
            # Basis-Zeile erstellen (erste Zeile der Gruppe als Grundlage)
            base_row = group_rows.iloc[0].copy()
            
            # Dictionary für die aggregierten Daten erstellen
            aggregated_data = {}
            
            # Alle Spalten der Basis-Zeile übernehmen
            for col in base_row.index:
                aggregated_data[col] = base_row[col]
            
            # Sammle alle verschiedenen Name 1 Varianten aus der Gruppe
            all_name1_variants = []
            for _, group_row in group_rows.iterrows():
                name1_value = group_row['Name 1']
                if pd.notna(name1_value) and name1_value not in all_name1_variants:
                    all_name1_variants.append(name1_value)
            
            # Setze den ersten Name 1 als Hauptname
            main_name = all_name1_variants[0] if all_name1_variants else None
            aggregated_data['Name 1'] = main_name
            
            # Füge andere Name 1 Varianten als Name 2, Name 3, etc. hinzu
            if len(all_name1_variants) > 1:
                print(f"  Verschiedene Name 1 Varianten gefunden: {all_name1_variants}")
                
                # Finde alle existierenden Name-Spalten und prüfe welche leer sind
                existing_name_cols = []
                for col in aggregated_data.keys():
                    if col.startswith('Name ') and col != 'Name 1':
                        match = re.match(r'Name (\d+)', col)
                        if match:
                            existing_name_cols.append((int(match.group(1)), col))
                
                # Sortiere nach Nummer
                existing_name_cols.sort()
                
                # Finde die nächste verfügbare Name-Spalte (leer oder neu)
                variant_index = 1  # Index für die zu verteilenden Varianten (überspringe erste)
                
                for variant in all_name1_variants[1:]:  # Überspringe den ersten (ist bereits Name 1)
                    added = False
                    
                    # Prüfe zuerst ob diese Variante bereits irgendwo existiert
                    already_exists = False
                    for existing_col in aggregated_data.keys():
                        if existing_col.startswith('Name '):
                            if pd.notna(aggregated_data[existing_col]) and aggregated_data[existing_col] == variant:
                                already_exists = True
                                print(f"    Name 1 Variante '{variant}' bereits vorhanden in {existing_col}")
                                break
                    
                    if not already_exists:
                        # Suche eine leere existierende Name-Spalte
                        for num, col_name in existing_name_cols:
                            if pd.isna(aggregated_data[col_name]) or aggregated_data[col_name] == "":
                                aggregated_data[col_name] = variant
                                print(f"    Name 1 Variante '{variant}' in leere Spalte {col_name} eingetragen")
                                added = True
                                break
                        
                        # Wenn keine leere Spalte gefunden, erstelle neue
                        if not added:
                            # Finde nächste freie Nummer
                            max_num = max([num for num, _ in existing_name_cols]) if existing_name_cols else 1
                            next_name_num = max_num + 1
                            new_name_col = f"Name {next_name_num}"
                            aggregated_data[new_name_col] = variant
                            print(f"    Name 1 Variante '{variant}' als neue Spalte {new_name_col} hinzugefügt")
            
            # Für jede weitere Zeile in der Gruppe die Daten hinzufügen
            for idx, (_, row) in enumerate(group_rows.iloc[1:].iterrows(), start=2):
                print(f"  Verarbeite Zeile {idx} aus Gruppe {group_values}...")
                
                # Statische Spalten übernehmen
                for static_col in static_columns:
                    if pd.notna(row[static_col]):
                        if static_col not in aggregated_data or pd.isna(aggregated_data[static_col]):
                            aggregated_data[static_col] = row[static_col]
                        elif static_col == 'Häufigkeit':
                            # Häufigkeiten addieren
                            try:
                                current_value = float(aggregated_data[static_col]) if pd.notna(aggregated_data[static_col]) else 0
                                new_value = float(row[static_col]) if pd.notna(row[static_col]) else 0
                                aggregated_data[static_col] = current_value + new_value
                                print(f"    Häufigkeit addiert: {current_value} + {new_value} = {aggregated_data[static_col]}")
                            except (ValueError, TypeError):
                                # Falls Häufigkeit nicht numerisch ist, behalte ersten Wert
                                print(f"    Warnung: Häufigkeit nicht numerisch: '{aggregated_data[static_col]}' vs '{row[static_col]}' - behalte ersten Wert")
                        elif aggregated_data[static_col] != row[static_col]:
                            print(f"    Warnung: Unterschiedliche Werte in {static_col}: '{aggregated_data[static_col]}' vs '{row[static_col]}' - behalte ersten Wert")
                
                # Spezielle Behandlung für Name-Spalten (prüfe auf Duplikate)
                name_cols = [col for col in result_df.columns if col.startswith('Name ') and col != 'Name 1']
                for name_col in name_cols:
                    if pd.notna(row[name_col]):
                        name_value = row[name_col]
                        
                        # Prüfe ob dieser Name bereits in den aggregierten Daten existiert (inklusive Name 1)
                        name_already_exists = False
                        for existing_col in aggregated_data.keys():
                            if existing_col.startswith('Name '):  # Prüfe alle Name-Spalten inklusive Name 1
                                if pd.notna(aggregated_data[existing_col]) and aggregated_data[existing_col] == name_value:
                                    name_already_exists = True
                                    print(f"    Name '{name_value}' bereits vorhanden in {existing_col}, überspringe Duplikat")
                                    break
                        
                        # Nur hinzufügen wenn Name noch nicht existiert
                        if not name_already_exists:
                            # Finde nächste freie Name-Nummer
                            name_numbers = []
                            for col in aggregated_data.keys():
                                if col.startswith('Name ') and col != 'Name 1':
                                    match = re.match(r'Name (\d+)', col)
                                    if match:
                                        name_numbers.append(int(match.group(1)))
                            
                            next_name_num = max(name_numbers) + 1 if name_numbers else 2
                            new_name_col = f"Name {next_name_num}"
                            aggregated_data[new_name_col] = name_value
                            print(f"    Name '{name_value}' als {new_name_col} hinzugefügt")
                
                # Für alle anderen nummerierten Spaltengruppen
                for base_name, col_list in numbered_patterns.items():
                    if base_name == 'Name':  # Name-Spalten wurden bereits behandelt
                        continue
                        
                    # Finde die nächste freie Nummer für diese Spaltengruppe
                    existing_numbers = []
                    for num, col_name in col_list:
                        if col_name in aggregated_data and pd.notna(aggregated_data[col_name]):
                            existing_numbers.append(num)
                    
                    # Bestimme die nächste verfügbare Nummer
                    next_num = max(existing_numbers) + 1 if existing_numbers else 1
                    
                    # Kopiere Daten aus der aktuellen Zeile
                    for orig_num, orig_col in col_list:
                        if pd.notna(row[orig_col]):  # Nur wenn Wert vorhanden
                            new_col_name = f"{base_name} {next_num}"
                            aggregated_data[new_col_name] = row[orig_col]
                            next_num += 1
            
            # Neue DataFrame erstellen: andere Zeilen + aggregierte Zeile
            all_columns = set(result_df.columns)
            all_columns.update(aggregated_data.keys())
            
            # Erstelle neue Zeile als Series
            new_row_data = {}
            for col in all_columns:
                if col in aggregated_data:
                    new_row_data[col] = aggregated_data[col]
                else:
                    new_row_data[col] = None
            
            # Füge die aggregierte Zeile hinzu
            new_row_df = pd.DataFrame([new_row_data])
            result_df = pd.concat([other_rows, new_row_df], ignore_index=True)
            
            print(f"  Aggregiert: {len(group_rows)} Zeilen zu 1 Zeile für Gruppe {group_values}")
        
        print(f"\n=== Aggregation abgeschlossen ===")
        print(f"Ursprüngliche Zeilen: {len(df)}")
        print(f"Finale Zeilen: {len(result_df)}")
        
        # Gelbe Zeilen für alle aggregierten Gruppen prüfen
        all_aggregated_values = [val for group in aggregation_groups for val in group]
        should_mark_yellow = False
        
        for row_idx in range(2, ws.max_row + 1):
            name1_cell = ws.cell(row=row_idx, column=name1_col_idx)
            if name1_cell.value in all_aggregated_values:
                # Prüfe ob diese Zeile gelb markiert ist
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if is_cell_yellow(cell):
                        should_mark_yellow = True
                        yellow_rows.add(row_idx)
                        break
        
        # Spalten in ursprünglicher Reihenfolge + neue Spalten am Ende
        original_columns = list(df.columns)
        new_columns = [col for col in result_df.columns if col not in original_columns]
        
        # Neue Spalten sortieren (Name-Spalten numerisch, andere alphabetisch)
        def sort_new_columns_key(col):
            if col.startswith('Name '):
                match = re.match(r'Name (\d+)', col)
                if match:
                    return (0, int(match.group(1)))  # Name-Spalten nach Nummer
            return (1, col)  # Andere neue Spalten alphabetisch
        
        new_columns_sorted = sorted(new_columns, key=sort_new_columns_key)
        
        # Finale Spaltenreihenfolge: Original + neue Spalten
        sorted_columns = original_columns + new_columns_sorted
        result_df = result_df[sorted_columns]
        
        # Neue Workbook erstellen mit Formatierung
        new_wb = Workbook()
        new_ws = new_wb.active
        
        # DataFrame in Workbook schreiben
        for r in dataframe_to_rows(result_df, index=False, header=True):
            new_ws.append(r)
        
        # Formatierung übertragen
        print("Übertrage Formatierung...")
        
        # Gelbe Füllfarbe definieren
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Formatierung für bestehende Zeilen übertragen
        # Mapping von ursprünglichen DataFrame-Indizes zu Excel-Zeilen erstellen
        original_to_excel_mapping = {}
        current_excel_row = 2  # Startet bei Zeile 2 (nach Header)
        
        for df_idx in range(len(df)):
            original_to_excel_mapping[df_idx] = current_excel_row
            current_excel_row += 1
        
        # Für jede Zeile in der neuen Datei
        for new_row_idx in range(2, len(result_df) + 2):  # +2 für Header und 1-basiert
            result_row_data = result_df.iloc[new_row_idx - 2]  # -2 für 0-basiert
            
            # Versuche die ursprüngliche Zeile zu finden
            original_row_found = False
            for orig_df_idx in range(len(df)):
                orig_row_data = df.iloc[orig_df_idx]
                
                # Prüfe ob dies die gleiche Zeile ist (vergleiche Name 1)
                if pd.notna(result_row_data['Name 1']) and pd.notna(orig_row_data['Name 1']):
                    if result_row_data['Name 1'] == orig_row_data['Name 1']:
                        # Prüfe weitere eindeutige Felder für bessere Identifikation
                        match_found = True
                        for check_col in ['Name 2', 'Name 3'] if 'Name 2' in df.columns else ['Name 1']:
                            if check_col in df.columns and check_col in result_df.columns:
                                if pd.notna(orig_row_data[check_col]) and pd.notna(result_row_data[check_col]):
                                    if orig_row_data[check_col] != result_row_data[check_col]:
                                        match_found = False
                                        break
                        
                        if match_found:
                            # Übertrage Formatierung von ursprünglicher Zeile
                            orig_excel_row = original_to_excel_mapping[orig_df_idx]
                            
                            for col_idx in range(1, len(sorted_columns) + 1):
                                if (orig_excel_row, col_idx) in yellow_cells:
                                    new_ws.cell(row=new_row_idx, column=col_idx).fill = yellow_fill
                            
                            original_row_found = True
                            break
            
            # Wenn keine ursprüngliche Zeile gefunden wurde, ist es eine aggregierte Zeile
            if not original_row_found and should_mark_yellow:
                # Prüfe ob diese aggregierte Zeile gelb markiert werden soll
                name1_value = result_row_data['Name 1']
                if pd.notna(name1_value) and name1_value in all_aggregated_values:
                    name1_col_idx_new = sorted_columns.index('Name 1') + 1
                    new_ws.cell(row=new_row_idx, column=name1_col_idx_new).fill = yellow_fill
                    print(f"Aggregierte Zeile für '{name1_value}' wurde gelb markiert")
        
        # Speichern
        new_wb.save(output_file)
        print(f"\nErgebnis gespeichert: {output_file}")
        print(f"Endergebnis: {len(result_df)} Zeilen, {len(result_df.columns)} Spalten")
        
        # Zusammenfassung der Aggregationen
        total_original_rows = len(df)
        total_aggregated_groups = sum(1 for group in aggregation_groups if len(df[df['Name 1'].isin(group)]) > 1)
        total_saved_rows = total_original_rows - len(result_df)
        print(f"Verarbeitet: {total_aggregated_groups} Aggregationsgruppen")
        print(f"Reduziert um: {total_saved_rows} Zeilen durch Aggregation")
        
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Datei: {e}")
        import traceback
        traceback.print_exc()

# Hauptfunktion
def main():
    input_file = "../datasheets/personen/Personen_Empfaenger_Aussteller_aggregiert_marked.xlsx"
    output_file = "../datasheets/personen/Personen_Empfaenger_Aussteller_aggregiert_marked_v2.xlsx"
    
    print("Starte Aggregation der Excel-Datei...")
    aggregate_excel_rows(input_file, output_file)
    print("Fertig!")

if __name__ == "__main__":
    main()