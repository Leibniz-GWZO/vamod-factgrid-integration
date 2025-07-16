import pandas as pd
from openpyxl import load_workbook

excel_path = "../datasheets/personen/Personen_Empfaenger_Aussteller_aggregiert_marked_v2_SJ.xlsx"

# Lade die Excel-Datei mit openpyxl, um Zellfarben zu lesen
wb = load_workbook(excel_path)
ws = wb.active

# Finde die Spaltennummer von "Normierte deutsche Schreibweise"
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
norm_col_idx = header.index("Normierte deutsche Schreibweise") + 1

# Sammle die Zeilennummern, deren "Normierte deutsche Schreibweise" grün markiert ist
green_rows = []
for row in ws.iter_rows(min_row=2):
    cell = row[norm_col_idx - 1]
    fill = cell.fill
    # Erweiterte Farberkennung für verschiedene Grüntöne
    if fill and fill.fgColor:
        if fill.fgColor.type == "rgb":
            rgb = str(fill.fgColor.rgb)
            if any(green_code in rgb for green_code in ["4EA72E", "00FF00", "92D050", "C6EFCE"]):
                green_rows.append(cell.row)
        elif fill.fgColor.type == "theme":
            green_rows.append(cell.row)
    # Zusätzlich prüfe auf Hintergrundfarbe
    elif fill and fill.bgColor:
        if fill.bgColor.type == "rgb":
            rgb = str(fill.bgColor.rgb)
            if any(green_code in rgb for green_code in ["4EA72E", "00FF00", "92D050", "C6EFCE"]):
                green_rows.append(cell.row)

print(f"Gefundene grün markierte Zeilen: {len(green_rows)}")

# Lese die Excel-Datei mit pandas
df = pd.read_excel(excel_path)

# Entferne "Lfd. Nr." Spalte falls vorhanden
if "Lfd. Nr." in df.columns:
    df = df.drop(columns=["Lfd. Nr."])

# Speichere die ursprüngliche Spaltenreihenfolge
original_columns = df.columns.tolist()

# Identifiziere grün markierte Zeilen im DataFrame
green_indices = [r-2 for r in green_rows]  # pandas index beginnt bei 0, Excel bei 2
df_green = df[df.index.isin(green_indices)]

print(f"Anzahl grün markierter Zeilen im DataFrame: {len(df_green)}")

# Erstelle eine Kopie des ursprünglichen DataFrames für das Resultat
result_df = df.copy()

# Finde Zeilen, die aggregiert werden sollen (grün markiert mit gleichem "Normierte deutsche Schreibweise")
to_aggregate = {}
to_remove_indices = []

for norm_name, group in df_green.groupby("Normierte deutsche Schreibweise"):
    if len(group) > 1:  # Nur aggregieren wenn mehr als eine Zeile
        print(f"Aggregiere Gruppe: {norm_name} mit {len(group)} Zeilen")
        
        # Erste Zeile als Basis nehmen
        base_row = group.iloc[0].copy()
        
        # Häufigkeit summieren
        base_row["Häufigkeit"] = group["Häufigkeit"].sum()
        
        # Definiere die Spaltentypen die aggregiert werden sollen
        column_types = ["Name", "Funktion", "Aussteller", "Datum Funktion", "Edition", "Rolle in Urkunde", "Geschlecht"]
        
        # Lösche alle vorhandenen Werte für die zu aggregierenden Spaltentypen
        for col_type in column_types:
            type_columns = [col for col in base_row.index if col == col_type or (col.startswith(col_type + " ") and len(col.split(" ")) == 2 and col.split(" ")[1].isdigit())]
            for col in type_columns:
                base_row[col] = None
        
        # Spezielle Behandlung für Name-Spalten (wie in der vorherigen Version)
        name_columns = [col for col in group.columns if col == "Name" or (col.startswith("Name ") and len(col.split(" ")) == 2 and col.split(" ")[1].isdigit())]
        all_name_values = []
        for _, row in group.iterrows():
            for col in name_columns:
                val = row[col]
                if pd.notna(val) and str(val) != 'nan' and str(val) != '':
                    all_name_values.append(str(val))
        
        # Entferne Duplikate für Namen, behalte Reihenfolge
        unique_name_values = []
        for v in all_name_values:
            if v not in unique_name_values:
                unique_name_values.append(v)
        
        # Verteile Namen auf nummerierte Spalten (alle ab 1 nummeriert)
        for i, val in enumerate(unique_name_values):
            col_name = f"Name {i+1}"
            
            # Erstelle Spalte falls sie nicht existiert
            if col_name not in result_df.columns:
                result_df[col_name] = None
            
            base_row[col_name] = val
        
        # Sammle alle Datensätze für andere Spaltentypen aus allen Zeilen (behalte Zusammenhänge)
        other_column_types = ["Funktion", "Aussteller", "Datum Funktion", "Edition", "Rolle in Urkunde", "Geschlecht"]
        all_data_sets = []
        
        for _, row in group.iterrows():
            # Finde die maximale Anzahl von Datensätzen in dieser Zeile
            max_num = 0
            for col_type in other_column_types:
                # Finde Spalten für diesen Typ in der aktuellen Zeile
                type_columns = [col for col in row.index if col == col_type or (col.startswith(col_type + " ") and len(col.split(" ")) > 1 and col.split(" ")[-1].isdigit())]
                for col in type_columns:
                    if " " in col:
                        try:
                            num = int(col.split(" ")[-1])
                            max_num = max(max_num, num)
                        except ValueError:
                            pass
                    else:
                        max_num = max(max_num, 1)

            # Sammle alle Datensätze aus dieser Zeile
            for i in range(1, max_num + 1):
                data_set = {}
                has_data = False
                
                for col_type in other_column_types:
                    if i == 1 and f"{col_type} 1" not in row.index:
                         # Handle base columns like "Funktion" which are equivalent to "Funktion 1"
                        col_name = col_type
                    else:
                        col_name = f"{col_type} {i}"
                    
                    if col_name in row.index:
                        val = row[col_name]
                        if pd.notna(val) and str(val) != 'nan' and str(val).strip() != '':
                            data_set[col_type] = str(val)
                            has_data = True
                
                if has_data:
                    all_data_sets.append(data_set)
        
        # Entferne Duplikate, um Redundanzen zu vermeiden
        unique_data_sets = []
        for data_set in all_data_sets:
            if data_set not in unique_data_sets:
                unique_data_sets.append(data_set)

        # Verteile die einzigartigen Datensätze auf nummerierte Spalten
        for i, data_set in enumerate(unique_data_sets):
            for col_type, val in data_set.items():
                col_name = f"{col_type} {i+1}"
                
                # Erstelle Spalte falls sie nicht existiert
                if col_name not in result_df.columns:
                    result_df[col_name] = None
                
                base_row[col_name] = val
        
        # Aktualisiere die erste Zeile der Gruppe
        result_df.loc[group.index[0]] = base_row
        
        # Markiere die anderen Zeilen zum Entfernen
        to_remove_indices.extend(group.index[1:].tolist())

# Entferne die aggregierten Zeilen
result_df = result_df.drop(index=to_remove_indices)

# Stelle sicher, dass die Spaltenreihenfolge erhalten bleibt und Name-Spalten gruppiert sind
name_cols = sorted([col for col in result_df.columns if col.startswith("Name ")], 
                   key=lambda x: int(x.split(" ")[1]))

other_cols = [col for col in original_columns if col in result_df.columns and not col.startswith("Name")]

# Füge neue Spalten hinzu, die nicht im Original waren (außer Name-Spalten)
new_other_cols = [col for col in result_df.columns if col not in original_columns and not col.startswith("Name ")]

# Finde die Position von "Normierte deutsche Schreibweise"
try:
    norm_idx = other_cols.index("Normierte deutsche Schreibweise")
    # Setze die Spalten zusammen: bis Normierte, dann alle Namen, dann der Rest
    final_columns = other_cols[:norm_idx + 1] + name_cols + other_cols[norm_idx + 1:] + new_other_cols
except ValueError:
    # Fallback, falls "Normierte deutsche Schreibweise" nicht gefunden wird
    final_columns = other_cols + name_cols + new_other_cols

result_df = result_df[final_columns]

print(f"Anzahl finaler Zeilen: {len(result_df)}")
print(f"Spalten in finaler Tabelle: {list(result_df.columns)}")

# Schreibe die aggregierte Tabelle als neue Excel-Datei
result_df.to_excel("../datasheets/personen/Personen_Empfaenger_Aussteller_final.xlsx", index=False)