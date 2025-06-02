import re
import pandas as pd
from openpyxl import load_workbook

# Pfad zur Eingabedatei
file_path = "../datasheets/180809_Urkunden_Personen_korrekt_Sortiert2.xlsx"

# Mapping: Blattname -> Header-Zeile (0-basiert)
header_mapping = {
    "EL": 2,
    "WO": 2,
    "LA": 2,
    "MU": 2,
    "CR": 2,
    "WII": 2,
    "Erzb.": 2,
    "JA": 1,
    "KIII": 0
}

# Aussteller-Benennungen
aussteller_mapping = {
    "KIII": "Kazimierz III.",
    "EL": "Ełżbieta Łokietkówna",
    "WO": "Władysław Opolczyk",
    "LA": "Ludwig von Anjou",
    "MU": "Maria von Ungarn",
    "CR": "Capitaneus Russie",
    "JA": "Jadwiga von Anjou",
    "WII": "Władysław II. Jagiełło",
}

# Spaltennamen für Bestätigungen je Blatt
sheet_best1 = {"KIII": "B1", "EL": "Bestätigung", "WO": "B1",
               "LA": "B1", "MU": "B1", "CR": "B1",
               "JA": "B1", "WII": "B1", "Erzb.": "Bestätigung 1"}
sheet_best2 = {"KIII": "B2", "EL": None, "WO": "B2",
               "LA": "B2", "MU": "B2", "CR": "B2",
               "JA": "B2", "WII": "B2", "Erzb.": "Bestätigung 2"}
sheet_best3 = {"KIII": "B3", "EL": None, "WO": "B3",
               "LA": "B3", "MU": "B3", "CR": "B3",
               "JA": "B3", "WII": "B3", "Erzb.": "Bestätigung 3"}

# Workbook laden, um Strikethrough zu erkennen
wb = load_workbook(filename=file_path, data_only=True)

# Hilfsfunktion für sichere String-Konvertierung
def get_str(row, col):
    val = row.get(col, "")
    if pd.isna(val):
        return ""
    return str(val).strip()

all_rows = []

for sheet_name, header_row_zero in header_mapping.items():
    header_excel = header_row_zero + 1
    ws = wb[sheet_name]

    # Strikethrough-Zeilen ermitteln
    striked = set()
    for row_cells in ws.iter_rows(min_row=header_excel+1):
        if any(cell.font and cell.font.strike for cell in row_cells):
            striked.add(row_cells[0].row)

    # DataFrame lesen
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=header_row_zero,
        engine="openpyxl",
        dtype=str
    )

    # Gestrichene Zeilen entfernen
    drop_idx = [r - header_excel - 1 for r in striked]
    df = df.drop(index=drop_idx, errors="ignore")

    # Leere Zeilen (außer "Nr.") entfernen
    if 'Nr.' in df.columns:
        cols_empty = [c for c in df.columns if c != 'Nr.']
        mask_empty = df[cols_empty].fillna('').applymap(lambda x: str(x).strip() == '').all(axis=1)
        df = df[~mask_empty]

    for _, row in df.iterrows():
        # Aussteller bestimmen
        if sheet_name == "Erzb.":
            rolle = get_str(row, "Rolle in Urkunde")
            if rolle not in ["Aussteller", "Aussteller 1", "Aussteller 2"]:
                continue
            aussteller = f"{get_str(row, 'Genannte Personen')}, {get_str(row, 'Funktion')}"
        else:
            aussteller = aussteller_mapping.get(sheet_name, "")

        # Datum vereinheitlichen
        if sheet_name not in ["EL", "LA", "MU", "Erzb."]:
            datum = get_str(row, "Jahr")
        else:
            year = get_str(row, "Jahr")
            dm = get_str(row, "Datum")
            day, month = "00", "00"
            parts = dm.split('.')
            if len(parts) >= 2:
                day = parts[0].strip().zfill(2)
                month = parts[1].strip().zfill(2)
            datum = f"{year}-{month}-{day}"

        # Bestätigungen
        b1 = get_str(row, sheet_best1[sheet_name]) if sheet_best1[sheet_name] else ""
        b2 = get_str(row, sheet_best2[sheet_name]) if sheet_best2[sheet_name] else ""
        b3 = get_str(row, sheet_best3[sheet_name]) if sheet_best3[sheet_name] else ""

        # Geschlecht/Familie
        if sheet_name == "KIII":
            geschlecht = get_str(row, "Geschlecht")
        elif sheet_name == "CR":
            geschlecht = get_str(row, "Familie")
        else:
            geschlecht = ""

        # Zeitabschnitte
        zeit_abschnitt = get_str(row, "Zeitabschnitt") if sheet_name in ["CR", "WII"] else ""
        zeit_abschnitt_cr = get_str(row, "Zeitabschnitt CR") if sheet_name == "WII" else ""

        all_rows.append({
            "Datum": datum,
            "Betreff": get_str(row, "Betreff"),
            "Ausstellungsort": get_str(row, "Ausstellungsort"),
            "Zielort": get_str(row, "Zielort"),
            "Genannte Personen": get_str(row, "Genannte Personen"),
            "Funktion": get_str(row, "Funktion"),
            "Rolle in Urkunde": get_str(row, "Rolle in Urkunde"),
            "Edition": get_str(row, "Edition"),
            "Anmerkung": get_str(row, "Anm."),
            "Bestätigung 1": b1,
            "Bestätigung 2": b2,
            "Bestätigung 3": b3,
            "Geschlecht": geschlecht,
            "Zeitabschnitt": zeit_abschnitt,
            "Zeitabschnitt CR": zeit_abschnitt_cr,
            "Aussteller": aussteller
        })

# Ergebnis-DataFrame
result_df = pd.DataFrame(all_rows)

# Datumsgenauigkeit bestimmen
safe_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
result_df['JahrExtrahiert'] = result_df['Datum'].str.extract(r"(\d{4})")[0]
mask_unsicher = ~result_df['Datum'].apply(lambda x: bool(safe_pattern.match(x)))
result_df['Datumsgenauigkeit'] = 'sicher'
result_df.loc[mask_unsicher, 'Datumsgenauigkeit'] = result_df.loc[mask_unsicher, 'Datum'].apply(lambda x: f"unsicher ({x})")
result_df.loc[mask_unsicher, 'Datum'] = result_df.loc[mask_unsicher, 'JahrExtrahiert'] + '-01-01'
result_df = result_df.drop(columns=['JahrExtrahiert'])

# Spaltenreihenfolge so anpassen, dass Datumsgenauigkeit direkt nach Datum kommt
cols = list(result_df.columns)
if 'Datum' in cols and 'Datumsgenauigkeit' in cols:
    cols.insert(cols.index('Datum') + 1, cols.pop(cols.index('Datumsgenauigkeit')))
result_df = result_df[cols]

# Pfad zur Ausgabedatei
output_path = "../datasheets/Urkunden_gesamt_neu.xlsx"
result_df.to_excel(output_path, index=False)
print(f"Fertig: '{output_path}' erstellt.")
