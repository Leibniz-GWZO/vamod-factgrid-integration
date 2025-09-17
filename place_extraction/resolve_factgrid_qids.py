import pandas as pd
import requests
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Excel-Datei einlesen
input_file = "../datasheets/orte/Orte_Identifikation_factgridS_SJ.xlsx"
df = pd.read_excel(input_file)

# 2. Spalte umbenennen: "Wikidata" → "Wikidata URL"
df = df.rename(columns={"Wikidata": "Wikidata URL"})

# 3. Wikidata-QID extrahieren
df["Wikidata QID"] = (
    df["Wikidata URL"]
    .dropna()
    .astype(str)
    .str.rstrip("/")
    .str.split("/")
    .str[-1]
)

# 4. Funktion zum Abrufen von FactGrid-URL und -QID
def get_factgrid_entry(qid: str):
    base = "https://database.factgrid.de/wiki/Special:ItemByTitle/wikidatawiki/{}"
    url = base.format(qid)
    resp = requests.get(url, allow_redirects=True)
    if resp.history:
        final_url = resp.url
        factgrid_qid = final_url.rstrip("/").split("/")[-1].split(":")[-1]
        return final_url, factgrid_qid
    else:
        return None, None

# 5. FactGrid-Infos abfragen mit tqdm für den Fortschritt
factgrid_urls = []
factgrid_qids = []

for q in tqdm(df["Wikidata QID"], desc="FactGrid-Abfragen"):
    if pd.isna(q) or not str(q).startswith("Q"):
        factgrid_urls.append(None)
        factgrid_qids.append(None)
    else:
        url, fg_qid = get_factgrid_entry(q)
        factgrid_urls.append(url)
        factgrid_qids.append(fg_qid)

# 6. Neue Spalten übernehmen
df["Factgrid URL"] = factgrid_urls
df["Factgrid QID"] = factgrid_qids

# 7. In eine neue Excel speichern mit Formatierung und allen Sheets
output_file = "../datasheets/orte/Orte_Identifikation_factgrid_v2.xlsx"

# Originale Arbeitsmappe laden
print("Loading original workbook with formatting...")
wb_original = load_workbook(input_file)

# Alle Sheets in neue Arbeitsmappe kopieren
wb_new = load_workbook(input_file)

# Das erste Sheet mit den neuen Daten aktualisieren
ws = wb_new.active
sheet_name = ws.title

# Alte Daten löschen (aber Formatierung beibehalten)
for row in ws.iter_rows():
    for cell in row:
        cell.value = None

# Neue Daten einfügen
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Arbeitsmappe speichern
wb_new.save(output_file)
print(f"Saved to {output_file} with preserved formatting and all sheets")

# 8. Ausgabe der ersten Zeilen (optional)
print(df[["Wikidata URL", "Wikidata QID", "Factgrid URL", "Factgrid QID"]].head())
